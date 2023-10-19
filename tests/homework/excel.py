from typing import Tuple
# 주의사항 1. 각 파일 간의 함수, 혹은 로직을 공유하지 마세요
# 주의사항 2. 함수 이름은 변경하지 마세요
# 주의사항 3. 함수명이 test_로 시작하는 함수를 만들지 마세요
# 위를 위반하면 채점이 제대로 되지 않아 0점 처리됩니다.


# 아래에서 pass를 지우고 로직을 작성하세요(10점)
# 아래 함수를 실행하면, assets.xlsx 에서 Corp.로 끝나는 회사를 다니는 사람들의 자산 총 합계를 리턴합니다.
# 리턴 값의 예) 10239102305
from openpyxl import load_workbook
def find_corp_total_asset() -> int:
    wb = load_workbook(filename = "/Users/besthersoy/yagomDE_2023/과제제출(github repository: script)/de-course-week4/hw_data/assets.xlsx")
    first_sheet = wb.active
    got_header  = False
    header = None
    data = []
    for r in first_sheet:
        if got_header is False:
            got_header  = True
            continue
        data.append({
            "name":r[0].value
           ,"company":r[1].value
           ,"city":r[2].value
           ,"age":r[3].value
           ,"est_asset_dollar":r[4].value
        })

    list = []
    for d in data:
        if d["company"][-5:] == "Corp.":
            list.append(d["est_asset_dollar"])

    list_sum = sum(list)
    return(list_sum)


# 아래에서 pass를 지우고 로직을 작성하세요(10점)
# 아래 함수를 실행하면, assets.xlsx 에서 LLC로 끝나는 회사를 다니는 사람들의 자산 총 평균을 리턴합니다.
# 리턴 값의 예) 1025831.87653
from openpyxl import load_workbook
def find_llc_total_asset() -> int:
    wb = load_workbook(filename = "/Users/besthersoy/yagomDE_2023/과제제출(github repository: script)/de-course-week4/hw_data/assets.xlsx")
    first_sheet = wb.active
    got_header  = False
    header = None
    data = []
    for r in first_sheet:
        if got_header is False:
            got_header  = True
            continue
        data.append({
            "name":r[0].value
           ,"company":r[1].value
           ,"city":r[2].value
           ,"age":r[3].value
           ,"est_asset_dollar":r[4].value
        })

    list = []
    for d in data:
        if d["company"][-3:] == "LLC":
            list.append(d["est_asset_dollar"])

    list_avg = sum(list) / len(list)
    return(list_avg)